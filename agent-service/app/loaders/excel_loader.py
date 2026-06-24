"""Excel 加载器 — 读取全部 Sheet，转为结构化文本"""

import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExcelLoader:
    """读取 .xlsx 所有 Sheet，提取表头 + 数据，输出结构化文本"""

    @staticmethod
    def supports(extension: str) -> bool:
        return extension.lower() in ("xlsx", "xls")

    def load(self, file_path: str) -> str:
        import openpyxl

        logger.info("[ExcelLoader] 开始解析: %s", file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        result_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result_parts.append(f"Sheet: {sheet_name}\n")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result_parts.append("(空表)\n")
                continue

            headers = [str(h) if h is not None else f"列{i+1}" for i, h in enumerate(rows[0])]
            result_parts.append("表头: " + " | ".join(headers))

            for row_idx, row in enumerate(rows[1:], start=2):
                values = []
                for col_idx, val in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
                    cell_val = str(val) if val is not None else ""
                    values.append(f"{header}: {cell_val}")
                result_parts.append(" | ".join(values))

            result_parts.append("")

        wb.close()
        text = "\n".join(result_parts)
        logger.info("[ExcelLoader] 解析完成, Sheet数=%d, 长度=%d", len(wb.sheetnames), len(text))
        return text